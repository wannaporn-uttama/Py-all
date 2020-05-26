from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "hello" #ชื่อ worksheet

ws['A1'] = "HelloTEST" #เพิ่มข้อมูลในช่อง A1

ws.append([5,6,7,8]) #เพิ่มข้อมูลหลายๆช่องพร้อมกัน
import datetime  
ws['A2'] = datetime.datetime.now() # เพิ่มวันเวลาในช่อง A2
wb.save("test.xlsx") # เซฟไฟล์
#wb = load_workbook(filename = 'test.xlsx') #test sample.xlsx

sheet_ranges = wb['hello'] #เรียกใช้ worksheet ที่ชื่อว่า hello

print(sheet_ranges['A1'].value) #อ่านค่าจากเซลส์ A1 ใน worksheet ของ hello